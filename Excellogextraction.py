#Excellogextraction.py - Extracts the log data from the excel file and saves it in an excel file
import openpyxl, os, logging, re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


#loop through folder to find all excel files

for excelFile in os.listdir('.'):
    if not excelFile.endswith('.xlsx'):
        continue
    wb = openpyxl.load_workbook(excelFile)
    sheet = wb.active
    #Create a new workbook
    wb2 = Workbook()
    sheet2 = wb2.active
    #Create the header row
    sheet2['A1'] = 'Date'
    sheet2['B1'] = 'Time'
    sheet2['C1'] = 'User'
    sheet2['D1'] = 'Computer'
    sheet2['E1'] = 'Description'
    #Find the last row in the sheet
    lastRow = sheet.max_row
    #Create a regex that matches the date and time
    dateRegex = re.compile(r'(\d{2}/\d{2}/\d{4})')
    timeRegex = re.compile(r'(\d{2}:\d{2}:\d{2})')
    #Loop through the rows
    for row in range(1, lastRow + 1):
        #Get the cell value
        cellValue = sheet['A' + str(row)].value
        #Search for the date and time
        date = dateRegex.search(cellValue)
        time = timeRegex.search(cellValue)
        #If a date and time is found, save it to the new sheet
        if date != None and time != None:
            sheet2['A' + str(row)] = date.group()
            sheet2['B' + str(row)] = time.group()
        #If a user is found, save it to the new sheet
        if 'User:' in cellValue:
            sheet2['C' + str(row)] = cellValue.split('User:')[1]
        #If a computer is found, save it to the new sheet
        if 'Computer:' in cellValue:
            sheet2['D' + str(row)] = cellValue.split('Computer:')[1]
        #If a description is found, save it to the new sheet
        if 'Description:' in cellValue:
            sheet2['E' + str(row)] = cellValue.split('Description:')[1]
    #Save the new workbook
    wb2.save(excelFile + '_log.xlsx')